"""FAQ sheet processor for product knowledge sheets.

Extracts Q&A pairs and descriptive content from the semi-structured
product sheets in the NUST Bank Excel file.
"""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl

from src.data_processing.base_processor import BaseProcessor, Document
from src.utils.logger import setup_logger

logger = setup_logger(__name__)


class FAQSheetProcessor(BaseProcessor):
    """Processes product FAQ sheets from the NUST Bank Excel file.

    Each sheet contains Q&A pairs and descriptive text about a bank product.
    The processor extracts these into Document objects with metadata.
    """

    def __init__(
        self,
        file_path: Path,
        sheet_names: list[str],
        skip_sheets: set[str] | None = None,
        rate_sheet_name: str = "Rate Sheet July 1 2024",
    ) -> None:
        self._file_path = file_path
        self._sheet_names = sheet_names
        self._skip_sheets = skip_sheets or set()
        self._rate_sheet_name = rate_sheet_name

    def process(self) -> list[Document]:
        """Process all FAQ sheets and return Document objects.

        Returns:
            List of Document objects with Q&A content and metadata.
        """
        wb = openpyxl.load_workbook(self._file_path, data_only=True)
        all_documents: list[Document] = []

        for sheet_name in self._sheet_names:
            if sheet_name in self._skip_sheets or sheet_name == self._rate_sheet_name:
                continue
            try:
                docs = self._process_sheet(wb, sheet_name)
                all_documents.extend(docs)
            except Exception as e:
                logger.error("Failed to process sheet '%s': %s", sheet_name, e)
                continue

        wb.close()
        logger.info(
            "FAQ sheets processed: %d documents from %d sheets",
            len(all_documents),
            len(self._sheet_names) - len(self._skip_sheets) - 1,
        )
        return all_documents

    def _process_sheet(
        self, wb: openpyxl.Workbook, sheet_name: str
    ) -> list[Document]:
        """Process a single FAQ sheet.

        Merges all consecutive non-question text blocks into the preceding
        Q&A answer so that feature lists, bullet points, and table rows are
        kept together with their parent question.  Standalone descriptive
        blocks that don't follow a question are also consolidated.
        """
        ws = wb[sheet_name]
        documents: list[Document] = []

        product_name = self._get_product_name(ws)
        text_blocks = self._extract_text_blocks(ws)

        i = 0
        while i < len(text_blocks):
            text = text_blocks[i]

            if self._is_question(text):
                # Consume ALL consecutive non-question blocks as the answer
                answer_parts: list[str] = []
                while (
                    i + 1 < len(text_blocks)
                    and not self._is_question(text_blocks[i + 1])
                ):
                    i += 1
                    answer_parts.append(text_blocks[i])

                answer = "\n".join(answer_parts)
                content = f"Q: {text}\nA: {answer}" if answer else f"Q: {text}"

                # Prefix with product name for disambiguation
                content = f"[{product_name}]\n{content}"
                documents.append(
                    Document(
                        content=content,
                        metadata={
                            "product": product_name,
                            "source_sheet": sheet_name,
                            "type": "qa",
                        },
                    )
                )
            else:
                # Standalone descriptive block — consolidate with following
                # non-question blocks until next question or end
                desc_parts = [text]
                while (
                    i + 1 < len(text_blocks)
                    and not self._is_question(text_blocks[i + 1])
                ):
                    i += 1
                    desc_parts.append(text_blocks[i])

                merged = "\n".join(desc_parts)
                if len(merged) > 30:
                    content = f"[{product_name}]\n{merged}"
                    documents.append(
                        Document(
                            content=content,
                            metadata={
                                "product": product_name,
                                "source_sheet": sheet_name,
                                "type": "description",
                            },
                        )
                    )
            i += 1

        return documents

    def _get_product_name(self, ws) -> str:
        """Extract product name from the first non-empty cell in the sheet."""
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            for cell_value in row:
                if cell_value and str(cell_value).strip():
                    name = str(cell_value).strip()
                    # Sometimes cell A1 has the product name (e.g., "Little Champs Account")
                    if len(name) > 3 and name != "Main":
                        return name
        return ws.title

    def _extract_text_blocks(self, ws) -> list[str]:
        """Extract all meaningful text blocks from the sheet.

        Merges multi-column content into single text blocks and filters
        out navigation links ('Main') and empty rows.
        """
        text_blocks: list[str] = []

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            # Collect non-empty cell values from this row
            parts = []
            for cell_value in row:
                if cell_value is not None:
                    text = str(cell_value).strip()
                    if text and text != "Main":
                        parts.append(text)

            if parts:
                combined = "\n".join(parts)
                if len(combined.strip()) > 0:
                    text_blocks.append(combined.strip())

        return text_blocks

    @staticmethod
    def _is_question(text: str) -> bool:
        """Determine if a text block is a question."""
        # Check if the first line ends with a question mark
        first_line = text.split("\n")[0].strip()
        return first_line.endswith("?")
