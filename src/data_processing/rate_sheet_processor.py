"""Semantic flattening processor for the Rate Sheet.

Converts the hierarchical, visually-structured rate sheet into
natural language sentences preserving full context for each rate.
"""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl

from src.data_processing.base_processor import BaseProcessor, Document
from src.utils.logger import setup_logger

logger = setup_logger(__name__)


class RateSheetProcessor(BaseProcessor):
    """Processes the 'Rate Sheet July 1 2024' sheet from the NUST Bank Excel file.

    Handles two regions (Savings Accounts left, Term Deposits right)
    and converts each rate row into a natural language sentence.
    """

    def __init__(
        self,
        file_path: Path,
        sheet_name: str = "Rate Sheet July 1 2024",
        effective_date: str = "July 1st, 2024",
    ) -> None:
        self._file_path = file_path
        self._sheet_name = sheet_name
        self._effective_date = effective_date

    def process(self) -> list[Document]:
        """Parse the rate sheet and produce semantic sentences.

        Returns:
            List of Document objects, one per rate entry.
        """
        wb = openpyxl.load_workbook(self._file_path, data_only=True)
        ws = wb[self._sheet_name]

        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            rows.append([cell.value for cell in row])

        documents: list[Document] = []
        documents.extend(self._process_savings(rows))
        documents.extend(self._process_term_deposits(rows))
        documents.extend(self._process_fcy(rows))

        wb.close()
        logger.info(
            "Rate sheet processed: %d documents from '%s'",
            len(documents),
            self._sheet_name,
        )
        return documents

    def _process_savings(self, rows: list[list]) -> list[Document]:
        """Extract savings account rates from columns B and D (indices 1 and 3)."""
        documents: list[Document] = []
        current_product = None

        for row in rows:
            label = self._clean(row[1]) if len(row) > 1 else ""  # col B
            rate  = self._clean(row[3]) if len(row) > 3 else ""  # col D

            if not label:
                continue

            # Skip section/global headers
            if "SAVINGS ACCOUNTS" in label or "Indicative Profit Rates" in label:
                continue

            # Skip sub-header row ("Profit Payment | Profit Rate")
            if "Profit Payment" in label:
                continue

            # Product name row: label present, no rate in col D
            if self._looks_like_product(label) and not self._is_rate(rate):
                current_product = label
                continue

            # Data row: label = payment frequency, rate = col D value
            if current_product and self._is_rate(rate):
                pct = self._to_percent(rate)
                sentence = (
                    f"The profit rate for the {current_product} is {pct}% per annum, "
                    f"with profit payments made {label}. "
                    f"This rate is effective from {self._effective_date}."
                )
                documents.append(
                    Document(
                        content=sentence,
                        metadata={
                            "product": current_product,
                            "category": "Savings Account",
                            "payment_frequency": label,
                            "rate": pct,
                            "source_sheet": self._sheet_name,
                            "effective_date": self._effective_date,
                        },
                    )
                )

        return documents

    def _process_term_deposits(self, rows: list[list]) -> list[Document]:
        """Extract term deposit rates from columns F, G, I (indices 5, 6, 8)."""
        documents: list[Document] = []
        current_product = None

        for row in rows:
            col_f = self._clean(row[5]) if len(row) > 5 else ""  # col F: header / product / tenor
            col_g = self._clean(row[6]) if len(row) > 6 else ""  # col G: payout
            col_i = self._clean(row[8]) if len(row) > 8 else ""  # col I: rate

            if not col_f and not col_g:
                continue

            # Stop when we reach the FCY section
            if col_f == "FCY":
                break

            # Skip the ALL-CAPS section header
            if col_f == "TERM DEPOSITS":
                continue

            # Skip sub-header rows (Tenor / Payout / Profit Rate column labels)
            if "Tenor" in col_f or col_f == "Payout":
                continue

            # Skip discontinued-product notices
            if "Discontinued" in col_f or "No Fresh Booking" in col_f:
                continue

            # Product name row: product keyword present, no rate in col I
            if self._looks_like_product(col_f) and not self._is_rate(col_i):
                current_product = col_f
                continue

            # Data row: col_f = tenor, col_g = payout, col_i = rate
            if current_product and col_f and self._is_rate(col_i):
                pct = self._to_percent(col_i)
                payout = col_g if col_g else "Maturity"
                sentence = (
                    f"The {current_product} for a tenor of {col_f} "
                    f"has a profit rate of {pct}% per annum, "
                    f"with payout at {payout}. "
                    f"This rate is effective from {self._effective_date}."
                )
                documents.append(
                    Document(
                        content=sentence,
                        metadata={
                            "product": current_product,
                            "category": "Term Deposit",
                            "tenor": col_f,
                            "payout": payout,
                            "rate": pct,
                            "source_sheet": self._sheet_name,
                            "effective_date": self._effective_date,
                        },
                    )
                )

        return documents

    def _process_fcy(self, rows: list[list]) -> list[Document]:
        """Extract foreign currency rates from the FCY section (col F onwards)."""
        documents: list[Document] = []

        for i, row in enumerate(rows):
            col_f = self._clean(row[5]) if len(row) > 5 else ""  # col F
            if col_f == "FCY":
                # Currency headers (USD, GBP, EUR…) start at col G (index 6)
                currencies = []
                for j in range(6, min(len(row), 10)):
                    c = self._clean(row[j])
                    if c:
                        currencies.append((j, c))

                # Following rows: col F = account type, col G/H/I = rates
                for data_row in rows[i + 1 :]:
                    acct = self._clean(data_row[5]) if len(data_row) > 5 else ""
                    if not acct:
                        break
                    for col_idx, currency in currencies:
                        rate = (
                            self._clean(data_row[col_idx])
                            if len(data_row) > col_idx
                            else ""
                        )
                        if self._is_rate(rate):
                            pct = self._to_percent(rate)
                            sentence = (
                                f"The profit rate for {currency} {acct} is "
                                f"{pct}% per annum. "
                                f"This rate is effective from {self._effective_date}."
                            )
                            documents.append(
                                Document(
                                    content=sentence,
                                    metadata={
                                        "product": f"{currency} {acct}",
                                        "category": "Foreign Currency",
                                        "currency": currency,
                                        "rate": pct,
                                        "source_sheet": self._sheet_name,
                                        "effective_date": self._effective_date,
                                    },
                                )
                            )
                break

        return documents

    @staticmethod
    def _clean(value) -> str:
        """Convert a cell value to a stripped string."""
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _is_rate(value) -> bool:
        """Check if a value looks like a decimal rate (0.0 to 1.0)."""
        if not value:
            return False
        try:
            f = float(value)
            return 0.0 <= f <= 1.0
        except (ValueError, TypeError):
            return False

    @staticmethod
    def _to_percent(value) -> str:
        """Convert a decimal rate to a formatted percentage string."""
        try:
            return f"{float(value) * 100:.2f}"
        except (ValueError, TypeError):
            return str(value)

    @staticmethod
    def _looks_like_product(text: str) -> bool:
        """Heuristic: does this text look like a product name?"""
        if not text:
            return False
        product_keywords = [
            "Account",
            "Deposit",
            "PLS",
            "Maximiser",
            "Waqaar",
            "Sahar",
            "Bachat",
            "Champs",
            "Remittance",
            "PakWatan",
            "Pensioner",
            "NUST",
            "Asaan",
            "Term Deposit",
            "SNDR",
        ]
        return any(kw.lower() in text.lower() for kw in product_keywords)
